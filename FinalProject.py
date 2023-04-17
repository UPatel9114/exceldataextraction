#importing openpyxl module
import openpyxl as oxl
workBook = oxl.load_workbook("Data.xlsx")
sheet = workBook.active
#------------------------------------------------------------------------------------------------------------------------------------------------#
#User-defined Function named employeeId()
#This function takes employee Id from the user and provides with the details as per user choice from the given 5 options.
def employeeId():
    flag=True #variable that holds a boolean value to control the while loop.
    print("\nSelect One option from below[1,2,3,4,5]:")
    options = [" 1.Employee data","2.Employee Contact info","3.Employee Address"
               ,"4.Employee name, salary and Department","5.Exit the Code Right now"] #Provided the five options for the user's choice.
    for i in range(1):
        print(options[i],'\t\t',options[i+1],'\n',options[i+2],'\t\t',options[i+3],'\n',options[i+4])
    option = input('Enter your choice:') #asking for the option choice
    #while loop that commands the whole part of the options provided.
    while flag:
        if option == '1': #This provide user the complete data of the employee whose id has been entered. the data include's ID,Name,Telephone,E-mail,Address,Department and Salary
            list1=[]
            employeeId = int(input('\nenter the employee id whose data you want[2 to 201]:'))
            if employeeId>201 or employeeId<2: #checking the range of the Id entered as it should not be excede the range.
                employeeId = int(input('\nenter the employee id again the valid range is 2 to 201'))
            for i in range(2,202):
                for j in range(1,8):
                    c=sheet.cell(employeeId,j)
                    list1.append(c.value) #import a whole rows data into list1 and then displays the data.
                #printing data using list indexing
                print('\nEmployee Id:',list1[0])
                print('Name:',list1[1])
                print('Contact[Telephone]:',list1[2])
                print('E-mail:',list1[3])
                print('Address:',list1[4])
                print('Department:',list1[5])
                print('Salary: $',list1[6])
                break #to break the for loop 
        elif option == '2': #This provide user the contact information of the employee whose id has been entered. the data include's Name,Telephone Number, and E-mail
            list1=[]
            employeeId = int(input('\nenter the employee id for the contact information[2 to 201]:'))
            if employeeId>201 or employeeId<2: #checking the range of the Id entered as it should not be excede the range.
                employeeId = int(input('\nenter the employee id again the valid range is 2 to 201'))
            for i in range(2,202):
                for j in range(1,8):
                    c=sheet.cell(employeeId,j)
                    list1.append(c.value) #import a whole rows data into list1 and then displays the data.
                print('\nName:',list1[1])
                print('Contact[Telephone]:',list1[2])
                print('E-mail:',list1[3])
                break
        elif option == '3': #This provide user the Address info of the employee whose id has been entered. the data include's Name and address of the employee.
            list1=[]
            employeeId = int(input('\nenter the employee id for the Address of the employee[2 to 201]:'))
            if employeeId>201 or employeeId<2: #checking the range of the Id entered as it should not be excede the range.
                employeeId = int(input('\nenter the employee id again the valid range is 2 to 201'))
            for i in range(2,202):
                for j in range(1,8):
                    c=sheet.cell(employeeId,j)
                    list1.append(c.value) #import a whole rows data into list1 and then displays the data.
                print('\nThe Name of the Employee is',list1[1],' who lives at ',list1[4])
                break
        elif option == '4': #This provide user the Salary and department data of the employee whose id has been entered. the data include's Name, department and salary of the employee.
            list1=[]
            employeeId = int(input('\nenter the employee id for salary and department info[2 to 201]:'))
            if employeeId>201 or employeeId<2: #checking the range of the Id entered as it should not be excede the range.
                employeeId = int(input('\nenter the employee id again the valid range is 2 to 201'))
            for i in range(2,202):
                for j in range(1,8):
                    c=sheet.cell(employeeId,j)
                    list1.append(c.value) #import a whole rows data into list1 and then displays the data.
                print('Name:',list1[1])
                print('Salary: $',list1[6])
                print('Department:',list1[5])
                break
        elif option == '5': #This option exit the code if user choices the option 5
            exit() #Exit function
        choice=input("\nDo you wish to continue:\nif Yes type y or Y\t\tif no type n or N") #asks the user if they wish to continue or exit the function.
        if choice == 'y' or choice == 'Y':
            flag = True
            print("Select One option from below[1,2,3,4,5]:")
            options = [" 1.Employee data","2.Employee Contact info","3.Employee Address"
               ,"4.Employee name, salary and department","5.Exit the Code Right now"]
            for i in range(1):
                print(options[i],'\t\t',options[i+1],'\n',options[i+2],'\t\t',options[i+3],'\n',options[i+4])
                option = input('Enter your choice:')
        elif choice == 'n' or choice == 'N':
            flag = False
            employee() # if false then goes back to employee() function.
        else:
            print("INVALID CHOICE!")
            choice=input("\nDo you wish to continue:\nif Yes type y or Y\t\tif no type n or N")
#--------------------------------------------------------------------------------------------------------------------------------------------#
#User-defined funtion named employeeName()
#This function takes a employee name from the user and gives the details as per user choice from the given 5 options.
def employeeName():
    flag=True
    print("\nSelect One option from below[1,2,3,4,5]:")
    options = [" 1.Employee data","2.Employee Contact info","3.Employee Address"
               ,"4.Employee name, salary and Department","5.Exit the Code Right now"]
    for i in range(1):
        print(options[i],'\t\t',options[i+1],'\n',options[i+2],'\t\t',options[i+3],'\n',options[i+4])
    option = input('Enter your choice:')
    columns = ["A","B","C","D","E","F","G"] #Defines the names of the columns in excel.
    #pre-defined lists to store the content from the columns in excel file.
    listId = []  #to store id's
    listName = [] #to store Name's
    listContact = [] #to store Contact's
    listEmail = [] #to store E-mails
    listAddress = [] #to store Addresses
    listDepartment = [] #to store Department's
    listSalary = [] #to store Salaries
    #for loops to append the data into the lists accordingly
    for i in columns[0]:
        for j in range(2,202):
            cell = sheet[str(i)+str(j)]
            listId.append(cell.value)
    for i in columns[1]:
        for j in range(2,202):
            cell = sheet[str(i)+str(j)]
            listName.append(cell.value)
    for i in columns[2]:
        for j in range(2,202):
            cell = sheet[str(i)+str(j)]
            listContact.append(cell.value)
    for i in columns[3]:
        for j in range(2,202):
            cell = sheet[str(i)+str(j)]
            listEmail.append(cell.value)
    for i in columns[4]:
        for j in range(2,202):
            cell = sheet[str(i)+str(j)]
            listAddress.append(cell.value)
    for i in columns[5]:
        for j in range(2,202):
            cell = sheet[str(i)+str(j)]
            listDepartment.append(cell.value)
    for i in columns[6]:
        for j in range(2,202):
            cell = sheet[str(i)+str(j)]
            listSalary.append(cell.value)
    employeeFullName = input("Enter the Full name of the Employee to access the data:") #asking for employee's full name
    #doing the input validation using while loop.
    while employeeFullName not in listName:
        employeeFullName = input("Enter the Full name of the Employee to access the data:")
    index = listName.index(employeeFullName) #store's the index value of the name provided
    #while to command the whole option part.
    while flag:
        if option == '1':
            print('\nEmployee Id:',listId[index])
            print('Name:',employeeFullName)
            print('Contact[Telephone]:',listContact[index])
            print('E-mail:',listEmail[index])
            print('Address:',listAddress[index])
            print('Department:',listDepartment[index])
            print('Salary: $',listSalary[index])
            break
        elif option == '2':
            print('Name:',employeeFullName)
            print('Contact[Telephone]:',listContact[index])
            print('E-mail:',listEmail[index])
            break
        elif option == '3':
            print('\nThe Name of the Employee is',employeeFullName,' who lives at ',listAddress[index])
            break
        elif option == '4':
            print('Name:',employeeFullName)
            print('Salary: $',listSalary[index])
            print('Department:',listDepartment[index])
            break
        elif option =='5':
            exit()
    employee() #goes back to employee() function.
#--------------------------------------------------------------------------------------------------------------------------------------#
#User-defined function named employeeDepartment()
#This function uses department name and provides with all the number of employee's who works in that department.
def employeeDepartment():
    columns = ["B","F"] #using only B and F columns as they represent the Department and name colunms in the excel file.
    #empty list's to store column data.
    listDepartment = []
    listName = []
    listOfEmployee = []
    #importing the data into the lists accordingly.
    for i in columns[0]:
        for j in range(2,202):
            cell = sheet[str(i)+str(j)]
            listName.append(cell.value)
    for i in columns[1]:
        for j in range(2,202):
            cell = sheet[str(i)+str(j)]
            listDepartment.append(cell.value)
    department = input("\nEnter the Department for the employees List:") #asking the user to enter a department of the company
    while department not in listDepartment: #input validation
        department = input("\nEnter the Department for the employees List:")
    for i in range(len(listDepartment)):
        if department == listDepartment[i]:
            listOfEmployee.append(listName[i])
    print("The list of employee's working in ",department,' department is as follows:')
    for i in range(len(listOfEmployee)):
        print(i+1,'.',listOfEmployee[i])
    choice = input("Do you want another department list or not[y for YES and n for NO]:") #asking if user wishes to continue using this function again.
    if choice == 'y' or choice == 'Y':
        employeeDepartment() #if yes then goes back to this function again
    elif choice == 'n' or choice == 'N':
        employee() #if no then goes back to employee function.
    else:
        choice = input("Do you want another department list or not[y for YES and n for NO]:")
#---------------------------------------------------------------------------------------------------------------------------------------------------#
#User-defined function named employee()
#This is the main function that provides the user with the option's to access the data accordingly.
def employee():
    print("\n1.Use employee ID to access data.")
    print("2.Use employee Name to access data.")
    print("3.Wants the Department wise employee List.")
    print("4.To exit the program right away.")
    choice = input("Enter the choice[1, 2, 3 or 4]:")
    while choice not in ['1','2','3','4']:
        choice = input("Enter the choice[1, 2, 3 or 4]:")
    if choice == '1':
        employeeId() #Leads to the employeeId function
    elif choice == '2':
        employeeName()#Leads to the employeeName function
    elif choice == '3':
        employeeDepartment()#Leads to the employeeDepartment function
    elif choice == '4':
        print()
employee() # calling employee function to start.
